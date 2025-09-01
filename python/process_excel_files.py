import argparse, json, re, string
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Please: pip install openpyxl")

# ---------------- Config ----------------
FILETYPE_NORMALIZE = {
    "audio": "Audio Bar",
    "audio bar": "Audio Bar",
    "embedded mp4": "Embedded MP4",
    "embedded  mp4": "Embedded MP4",
    "mp4": "Embedded MP4",
    "video": "Embedded MP4",
    "embedded mp3": "Embedded MP3",
    "mp3": "Embedded MP3",
}


# --------------- Helpers ----------------
def _canon(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = " ".join(s.split())
    return "".join(ch for ch in s if ch not in set(string.punctuation + " "))


def clean(s):
    if s is None:
        return None
    s = str(s).strip()
    return " ".join(s.split())


def norm_filetype(raw):
    if raw is None or str(raw).strip() == "":
        return "Audio Bar"
    key = " ".join(str(raw).strip().lower().split())
    return FILETYPE_NORMALIZE.get(key, str(raw).strip())


# Robust sheet-name parser (Level/Unit/Lesson anywhere in the title)
SHEET_RE = re.compile(
    r"""(?ix)
    (?:^|[^0-9a-z])
    (?:lvl|level|l)\s*[_\-\s]*?(?P<level>\d+)
    .*?
    (?:unt|unit|u)\s*[_\-\s]*?(?P<unit>\d+)
    .*?
    (?:lsn|lesson|les|ls)\s*[_\-\s]*?(?P<lesson>\d+)
    (?:[^0-9a-z]|$)
    """
)


def parse_sheet_signature(sheet_name: str):
    name = sheet_name.strip()
    m = SHEET_RE.search(name)
    if m:
        L = int(m.group("level"))
        U = int(m.group("unit"))
        S = int(m.group("lesson"))
        return f"level_{L}_unit_{U}_lesson_{S}", L, U, S
    nums = re.findall(r"\d+", name)
    if len(nums) >= 3:
        L, U, S = map(int, nums[:3])
        return f"level_{L}_unit_{U}_lesson_{S}", L, U, S
    squished = re.sub(r"[^0-9a-z]+", "_", name.lower())
    m2 = SHEET_RE.search(squished)
    if m2:
        L = int(m2.group("level"))
        U = int(m2.group("unit"))
        S = int(m2.group("lesson"))
        return f"level_{L}_unit_{U}_lesson_{S}", L, U, S
    raise ValueError(f"Cannot parse Level/Unit/Lesson from sheet name: {sheet_name!r}")


# Explicit header mapping that disambiguates "Audio #" vs "AUDIO"
def map_header(raw):
    if raw is None:
        return None
    text_raw = str(raw).strip()
    text = text_raw.lower()

    # ---- ignore counters & aggregates ----
    if re.fullmatch(r"audio\s*#", text):
        return None
    if re.fullmatch(r"#\s*audios", text) or text == "audios":
        return None
    if re.fullmatch(r"(#\s*)?slides?", text):
        return "slide_number"

    # ---- positive mappings ----
    if text in ("slide number", "slidenumber", "slide"):
        return "slide_number"
    if text in ("lesson#", "lesson"):
        return "lesson"
    if "block" in text and "section" in text:
        return "section"
    if "assignable" in text and "unit" in text:
        return "title"
    if text in ("script", "transcription", "audio"):
        return "transcription"
    if "file location" in text and "html" in text:
        return "file_type"
    if text in ("file type", "filetype", "file_type"):
        return "file_type"
    if text.startswith("notes"):
        return "notes"
    if text in (
        "occurrence",
        "occurence",
        "slide title",
        "slidetitle",
        "link to dev",
        "linktodev",
    ):
        return None
    return None


def find_header_row(ws, max_scan=8):
    hi = min(ws.max_row, max_scan)
    for r in range(1, hi + 1):
        row_vals = [cell.value for cell in ws[r]]
        mapped = [map_header(v) for v in row_vals]
        if "slide_number" in mapped and "transcription" in mapped:
            return r, row_vals
    return 2, [cell.value for cell in ws[2]]


def header_map(ws):
    hdr_row, hdr_values = find_header_row(ws)
    col_idx = {}
    for idx, h in enumerate(hdr_values):
        m = map_header(h)
        if m and m not in col_idx:
            col_idx[m] = idx
    return col_idx, hdr_row


# ---------------- Codes Loader ----------------
def _load_codes():
    codes_path = Path(__file__).parent / "data" / "codes.json"
    if not codes_path.exists():
        return {}
    try:
        with codes_path.open("r", encoding="utf-8") as f:
            raw = json.load(f) or {}
    except Exception:
        return {}

    def _to_base_key(name: str) -> str:
        try:
            base, _, _, _ = parse_sheet_signature(name)
            return base
        except Exception:
            s = re.sub(r"[^0-9a-z]+", "_", str(name).strip().lower())
            return s.strip("_")

    norm = {}
    for k, v in raw.items():
        try:
            norm[_to_base_key(k)] = v
        except Exception:
            continue
    return norm


_CODES_BY_BASE = _load_codes()


# --------------- Core ----------------
def excel_to_json_grouped(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out_sheets = []
    grand_total_slides = 0
    grand_total_audio = 0

    for ws in wb.worksheets:
        try:
            base, level, unit, lesson_num = parse_sheet_signature(ws.title)
        except Exception as e:
            raise ValueError(
                f"{xlsx_path.name}: cannot parse Level/Unit/Lesson from sheet name: {ws.title!r}"
            ) from e

        cols, header_row = header_map(ws)

        needed = ["slide_number", "transcription"]
        missing = [c for c in needed if c not in cols]
        if missing:
            raise ValueError(
                f"Sheet {ws.title!r} missing columns: {missing} (header row={header_row})"
            )

        last_slide = None
        last_section = None
        last_title = None

        audio_counts = {}
        slides_by_num = {}

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):

            def get(colname):
                return row[cols[colname]] if colname in cols else None

            slide = get("slide_number")
            section = get("section")
            title = get("title")
            transcript = get("transcription")
            notes = get("notes")
            ftype = get("file_type")

            if slide is None or str(slide).strip() == "":
                slide = last_slide
            else:
                try:
                    slide = int(str(slide).strip())
                except Exception:
                    continue
                last_slide = slide
                audio_counts.setdefault(slide, 0)

            if slide is None:
                continue

            section = clean(section) or last_section
            title = clean(title) or last_title
            last_section, last_title = section, title

            if (
                transcript is None
                and notes is None
                and (ftype is None or str(ftype).strip() == "")
            ):
                continue

            if slide not in slides_by_num:
                slides_by_num[slide] = {
                    "slide_number": slide,
                    "lesson": f"Lesson {lesson_num}",
                    "section": section,
                    "title": title,
                    "audio": [],
                }

            audio_counts[slide] = audio_counts.get(slide, 0) + 1
            a_idx = audio_counts[slide]

            audio_item = {
                "filename": f"{base}_{slide}_{a_idx}",
                "file_type": norm_filetype(ftype),
                "transcription": clean(transcript),
            }
            n = clean(notes)
            if n:
                audio_item["notes"] = n

            slides_by_num[slide]["audio"].append(audio_item)

        slides = []
        total_audio_this_sheet = 0
        for s in sorted(slides_by_num.values(), key=lambda x: x["slide_number"]):
            s["audio_total"] = len(s["audio"])
            total_audio_this_sheet += s["audio_total"]
            slides.append(s)

        toc = []
        for s in slides:
            short_title = re.sub(
                r"^Slide\s*\d+\s*:\s*", "", s.get("title") or "", flags=re.IGNORECASE
            ).strip()
            toc.append(
                {
                    "slide_number": s["slide_number"],
                    "section": s.get("section"),
                    "title": short_title or (s.get("title") or ""),
                }
            )

        sheet_obj = {
            "sheet_name": ws.title,
            "base": base,
            "level": level,
            "unit": unit,
            "lesson_num": lesson_num,
            "toc": toc,
            "slides": slides,
            "totals": {"slides": len(slides), "audio_items": total_audio_this_sheet},
        }

        if base in _CODES_BY_BASE:
            sheet_obj["xcode"] = _CODES_BY_BASE[base]

        out_sheets.append(sheet_obj)
        grand_total_slides += len(slides)
        grand_total_audio += total_audio_this_sheet

    out_sheets.sort(key=lambda s: (s["unit"], s["lesson_num"], s["sheet_name"]))

    return {
        "workbook": Path(xlsx_path).name,
        "totals": {
            "sheets": len(out_sheets),
            "slides": grand_total_slides,
            "audio_items": grand_total_audio,
        },
        "sheets": out_sheets,
    }


# --------------- CLI ----------------
def main():
    ap = argparse.ArgumentParser(
        description="Excel → JSON (grouped by sheet; filenames from sheet name)"
    )

    ap.add_argument(
        "-i",
        "--input",
        required=True,
        type=Path,
        nargs="+",
        help="Input .xlsx file(s) or directory containing .xlsx files",
    )
    ap.add_argument(
        "-o",
        "--output",
        required=True,
        type=Path,
        help="Output directory to write JSON files",
    )
    ap.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = ap.parse_args()

    # Collect .xlsx files from all inputs (files and/or directories)
    xlsx_files = []
    for path in args.input:
        if path.is_dir():
            xlsx_files.extend(sorted(path.glob("*.xlsx")))
        elif path.is_file() and path.suffix == ".xlsx":
            xlsx_files.append(path)
        else:
            print(f"Skipping invalid input: {path}")

    if not xlsx_files:
        raise SystemExit("No valid .xlsx files found to process.")

    # Ensure output directory exists
    args.output.mkdir(parents=True, exist_ok=True)

    for input_path in xlsx_files:
        try:
            data = excel_to_json_grouped(input_path)
            out_filename = input_path.with_suffix(".json").name
            out_path = args.output / out_filename

            with out_path.open("w", encoding="utf-8") as f:
                json.dump(
                    data, f, ensure_ascii=False, indent=2 if args.pretty else None
                )

            print(
                f"Wrote {out_path}  sheets={data['totals']['sheets']}  "
                f"slides={data['totals']['slides']}  audio={data['totals']['audio_items']}"
            )
        except Exception as e:
            print(f"❌ Error processing {input_path} — {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()


# python process_excel_files.py \
#   -i excel \
#   -o outputs \
#   --pretty
